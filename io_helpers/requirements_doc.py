# io_helpers/requirements_doc.py
import os
from pathlib import Path
import numpy as np
import pandas as pd
from io_helpers.find_product_folder import find_config
from openpyxl import load_workbook
import zipfile
import warnings  
warnings.filterwarnings("ignore", message="Data Validation extension is not supported and will be removed")  


def is_valid_xlsx(path):
    try:
        with zipfile.ZipFile(path, 'r') as zipf:
            return "[Content_Types].xml" in zipf.namelist()
    except zipfile.BadZipFile:
        return False


def safe_filename_token(value, fallback="scan"):
    text = str(value or "").strip()
    safe = "".join(ch if ch.isalnum() or ch in "-_." else "_" for ch in text)
    safe = safe.strip("._")
    return (safe or fallback)[:80]


def find_product_spec_v2(product_dir):
    product_path = Path(product_dir)

    direct = product_path / "ProductSpecifications_V2.xlsx"
    if direct.is_file():
        return direct

    child_matches = sorted(product_path.glob("*/ProductSpecifications_V2.xlsx"))
    if child_matches:
        return child_matches[0]

    nested_matches = sorted(product_path.rglob("ProductSpecifications_V2.xlsx"))
    if nested_matches:
        return nested_matches[0]

    raise FileNotFoundError(
        f"No valid ProductSpecifications_V2.xlsx found under: {product_path}"
    )


def resolve_product_spec_from_scan(scan_text):
    text = str(scan_text or "").strip()
    if not text:
        raise ValueError("Empty QR code; cannot resolve product specification.")

    scanned_path = Path(os.path.expanduser(text))
    if scanned_path.exists():
        scanned_path = scanned_path.resolve()
        product_dir = scanned_path.parent if scanned_path.is_file() else scanned_path
    else:
        old_cwd = os.getcwd()
        try:
            product_dir_raw = find_config(text)
        finally:
            try:
                os.chdir(old_cwd)
            except Exception:
                pass
        if not product_dir_raw:
            raise FileNotFoundError(f"[ERROR] No matching folder found for QR code: {text}")
        product_dir = Path(product_dir_raw)

    spec_path = find_product_spec_v2(product_dir)
    return spec_path.parent, spec_path


class ConfigReader:
    """Reads and stores configuration parameters and recepies from Excel files dynamically."""

    def __init__(self, qr_code):
        product_dir, spec_path = resolve_product_spec_from_scan(qr_code)

        # Dynamically set base path for this QR code.
        self.path = os.fspath(product_dir)
        self.BASE_PATH = self.path
        self.RECEPIES_FOLDER = os.path.join(self.BASE_PATH, "02_EOL_Recepies")
        self.PRODUCT_SPEC_V2 = os.fspath(spec_path)
        self.PRODUCT_SPEC = os.path.join(self.BASE_PATH, "ProductSpecifications.xlsx")
        self.STEERING_APP_TEST_FILE = os.path.join(self.RECEPIES_FOLDER, "SteeringAppRecepie.xlsx")
        self.STEERING_MOTOR_TEST_FILE = os.path.join(self.RECEPIES_FOLDER, "SteeringMotorRecepie.xlsx")
        self.TRACTION_TEST_FILE = os.path.join(self.RECEPIES_FOLDER, "TractionRecepie.xlsx")
        self.SAVE_RESULTS_PATH = os.path.join(self.BASE_PATH, "04_Results")
        os.makedirs(self.SAVE_RESULTS_PATH, exist_ok=True)

        # Use the same workbook as CalibrationGUI. EOL appends its sheets to
        # <serial>_status.xlsx and renames it to <serial>_done.xlsx on success.
        result_token = safe_filename_token(qr_code, 'qr')
        status_path = os.path.join(self.SAVE_RESULTS_PATH, f"{result_token}_status.xlsx")
        done_path = os.path.join(self.SAVE_RESULTS_PATH, f"{result_token}_done.xlsx")
        self.results_excel_path = (
            status_path
            if os.path.exists(status_path) or not os.path.exists(done_path)
            else done_path
        )

        if not os.path.exists(self.results_excel_path):
            with pd.ExcelWriter(self.results_excel_path, engine='openpyxl', mode='w') as writer:
                pd.DataFrame({'info': ['initialized']}).to_excel(writer, sheet_name='init', index=False)

        # Load config
        self.parameters = {}
        self.recepies = {}
        self.load_config()
        self.load_recepies()

    def load_config(self):
        """Loads configuration from the available Excel file dynamically."""
        if os.path.exists(self.PRODUCT_SPEC_V2) and is_valid_xlsx(self.PRODUCT_SPEC_V2):
            print(f"[INFO] Loading product specification: {self.PRODUCT_SPEC_V2}")
            self._read_spec(self.PRODUCT_SPEC_V2)
        else:
            raise FileNotFoundError(f"No valid product specification file found: {self.PRODUCT_SPEC_V2}")

    def _read_spec(self, filepath):
        """Reads all parameters from the available specification file dynamically."""
        xls = pd.ExcelFile(filepath, engine="openpyxl")
        for sheet in xls.sheet_names:
            df = pd.read_excel(filepath, sheet_name=sheet, engine="openpyxl")
            for index, row in df.iterrows():
                key = row.iloc[0]
                value = row.iloc[1]
                self.parameters[key] = value

    def get_product_parameter(self, name, default=None):
        """Retrieve a parameter by name safely."""
        return self.parameters.get(name, default)

    def load_recepies(self):
        """Loads recepies from additional Excel files."""
        for file_path, key in [
            (self.STEERING_APP_TEST_FILE, "SteeringApp"),
            (self.STEERING_MOTOR_TEST_FILE, "Steering"),
            (self.TRACTION_TEST_FILE, "Traction")
        ]:
            if os.path.exists(file_path):
                self.recepies[key] = self._read_recepies(file_path)

    def _read_recepies(self, filepath):
        """Reads only 'Sheet1' from the recepie file and stores it as a DataFrame."""
        try:
            df = pd.read_excel(filepath, sheet_name="Sheet1", engine="openpyxl")
            return df
        except ValueError:
            print(f"[ERROR] Recepies not found in {filepath}")
            return None

    # Save the results
    def save_results_to_excel(self, results, sheet_name):
        # Step 1: Save the new results first
        with pd.ExcelWriter(self.results_excel_path, engine='openpyxl', mode='a', if_sheet_exists='replace') as writer:
            results.to_excel(writer, sheet_name=sheet_name, index=False)

        # Step 2: Remove 'init' sheet if it still exists
        book = load_workbook(self.results_excel_path)
        if 'init' in book.sheetnames and sheet_name != 'init':
            del book['init']
            book.save(self.results_excel_path)

    @staticmethod
    def _match_value_is_blank(value):
        if pd.isna(value):
            return True
        return str(value).strip() == ""

    @staticmethod
    def _match_value_is_true(value):
        if isinstance(value, (bool, np.bool_)):
            return bool(value)
        if isinstance(value, (int, float)) and not pd.isna(value):
            return float(value) == 1.0
        return str(value).strip().lower() in {"true", "1", "yes", "ok"}

    def general_sheets_all_match(self):
        try:
            xls = pd.ExcelFile(self.results_excel_path, engine="openpyxl")
        except FileNotFoundError:
            return False

        general_sheets = [sheet for sheet in xls.sheet_names if sheet.startswith("general")]
        if not general_sheets:
            return False

        checked_any = False
        for sheet in general_sheets:
            df = pd.read_excel(self.results_excel_path, sheet_name=sheet, engine="openpyxl")
            if "Match" not in df.columns:
                return False
            for value in df["Match"].tolist():
                if self._match_value_is_blank(value):
                    continue
                checked_any = True
                if not self._match_value_is_true(value):
                    return False
        return checked_any

    def rename_status_to_done_if_eol_ok(self):
        if not self.general_sheets_all_match():
            return False

        current_path = Path(self.results_excel_path)
        if current_path.stem.endswith("_done"):
            return True
        if not current_path.stem.endswith("_status"):
            return False

        done_path = current_path.with_name(
            f"{current_path.stem[:-len('_status')]}_done{current_path.suffix}"
        )
        current_path.replace(done_path)
        self.results_excel_path = os.fspath(done_path)
        print(f"[INFO] EOL successful. Result workbook renamed to: {done_path}")
        return True


    def extract_brake_feedback(self, dp, pg, brake_torque):
        fragments = []
        try:
            workbook_sheets = set(pd.ExcelFile(self.results_excel_path, engine="openpyxl").sheet_names)
        except FileNotFoundError:
            print(f"[WARNING] Cannot build brake report; workbook not found: {self.results_excel_path}")
            return fragments

        for idx, sheet_name in enumerate(["brake_test_forward", "brake_test_backward"]):
            if sheet_name not in workbook_sheets:
                print(f"[INFO] Skipping brake report section {sheet_name}: sheet was not produced in this run.")
                continue
            try:
                actual_df = pd.read_excel(self.results_excel_path, sheet_name=sheet_name, engine="openpyxl")
                fig_parts = []

                include_js = 'cdn' if idx == 0 else False

                # Braking current (only when brake is active)
                brake_current = actual_df.loc[actual_df["brake activated"] == 1, "RMS current"]
                avg_current = brake_current.mean()
                Kt_estimated = brake_torque / avg_current if avg_current != 0 else 0.1

                # --- RMS Current + Brake Activation ---
                fig_current = pg.plot_current_with_brake(
                    time_list=actual_df['time(S)'].values,
                    current=actual_df['RMS current'].values,
                    brake_signal=actual_df['brake activated'].astype(int).values,
                    title=f"{sheet_name.replace('_', ' ').title()} - Current"
                )
                fig_parts.append(fig_current.to_html(full_html=False, include_plotlyjs=include_js))

                # --- Traction Feedback + Brake Activation ---
                fig_speed = pg.plot_speed_with_brake(
                    time_list=actual_df['time(S)'].values,
                    speed=actual_df['Traction feedback'].values,
                    brake_signal=actual_df['brake activated'].astype(int).values,
                    title=f"{sheet_name.replace('_', ' ').title()} - Speed"
                )
                fig_parts.append(fig_speed.to_html(full_html=False, include_plotlyjs=include_js))

                # --- Brake Energy ---
                fig_energy = pg.plot_energy_comparison(
                    time_list=actual_df["time(S)"].values,
                    current_list=actual_df["RMS current"].values,
                    speed_rpm_list=actual_df["Traction feedback"].values,
                    voltage=48,
                    kt=Kt_estimated,
                    title=f"{sheet_name.replace('_', ' ').title()} - Energy Comparison"
                )
                fig_parts.append(fig_energy.to_html(full_html=False, include_plotlyjs=include_js))

                fragments.append({
                    "title": sheet_name.replace("_", " ").title(),
                    "figs": fig_parts
                })
            except (ValueError, FileNotFoundError, KeyError) as exc:
                print(f"[INFO] Skipping brake report section {sheet_name}: {exc}")
        return fragments

    def extract_motor_combined_feedback(self, dp, pg):
        result_sections = []

        motor_tests = [
            {
                "title": "Traction",
                "sheet": "traction_test",
                "csv_pattern": "*_traction_motor_test.csv",
                "rename": {},
                "type": "feedback",
            },
            {
                "title": "Steering",
                "sheet": "steering_test",
                "csv_pattern": "*_steering_motor_test.csv",
                "rename": {
                    'speed_p': 'set point',
                    'feedback_speed': 'Traction feedback'
                },
                "type": "feedback",
            },
            {
                "title": "Steering App",
                "sheet": "steering_app",
                "csv_pattern": "*_steering_app_test.csv",
                "rename": {},
                "type": "position",
            }
        ]

        try:
            workbook_sheets = set(pd.ExcelFile(self.results_excel_path, engine="openpyxl").sheet_names)
        except FileNotFoundError:
            print(f"[WARNING] Cannot build motor report; workbook not found: {self.results_excel_path}")
            return result_sections

        def has_numeric_data(df, column):
            return column in df.columns and pd.to_numeric(df[column], errors="coerce").notna().any()

        def raw_pairs(df, value_column):
            return df[['time(S)', value_column]].values.tolist()

        for test in motor_tests:
            sheet = test["sheet"]
            title = test["title"]
            rename_map = test["rename"]
            test_type = test["type"]

            if sheet not in workbook_sheets:
                print(f"[INFO] Skipping {title} report section: sheet '{sheet}' was not produced in this run.")
                continue

            try:
                actual_df = pd.read_excel(self.results_excel_path, sheet_name=sheet, engine="openpyxl")
                if rename_map:
                    actual_df = actual_df.rename(columns=rename_map)

                try:
                    _all_df, avg_df = dp.all_files(
                        self.SAVE_RESULTS_PATH,
                        test["csv_pattern"],
                        sheet,
                        os.path.basename(self.results_excel_path)
                    )
                    if rename_map:
                        avg_df = avg_df.rename(columns=rename_map)
                except ValueError:
                    print(f"[INFO] No comparison history for {title}; plotting current run only.")
                    avg_df = actual_df.copy()

                fig_parts = []

                if test_type == "feedback":
                    required = {'time(S)', 'set point', 'Traction feedback'}
                    missing_required = sorted(required - set(actual_df.columns))
                    if missing_required:
                        raise KeyError(f"missing required columns: {missing_required}")

                    data_dict = {
                        "Speed RPM": (
                            raw_pairs(actual_df, 'Traction feedback'),
                            dp.min_max_avg(actual_df, 'Traction feedback'),
                            raw_pairs(avg_df, 'Traction feedback'),
                            dp.min_max_avg(avg_df, 'Traction feedback'),
                        )
                    }

                    optional_columns = [
                        ("DC Motor Current", "current A"),
                        ("RMS Current", "RMS current"),
                        ("Temperature", "temperature"),
                    ]
                    for param_name, column in optional_columns:
                        if has_numeric_data(actual_df, column) and has_numeric_data(avg_df, column):
                            data_dict[param_name] = (
                                raw_pairs(actual_df, column),
                                dp.min_max_avg(actual_df, column),
                                raw_pairs(avg_df, column),
                                dp.min_max_avg(avg_df, column),
                            )
                        else:
                            print(f"[INFO] Skipping {title} {param_name} plot: no numeric '{column}' data.")

                    for param_name, data in data_dict.items():
                        fig = pg.plot_motor_feedback(param_name, data)
                        include_js = 'cdn' if not fig_parts else False
                        fig_parts.append(fig.to_html(full_html=False, include_plotlyjs=include_js))

                elif test_type == "position":
                    required = {'time(S)', 'position input', 'position feedback'}
                    missing_required = sorted(required - set(actual_df.columns))
                    if missing_required:
                        raise KeyError(f"missing required columns: {missing_required}")

                    raw_pos_input = raw_pairs(actual_df, 'position input')
                    raw_pos_feedback = raw_pairs(actual_df, 'position feedback')
                    gen_pos_feedback = raw_pairs(avg_df, 'position feedback')

                    fig_parts.append(
                        pg.plot_motor_app_feedback(
                            "Position Feedback",
                            raw_pos_feedback,
                            gen_pos_feedback,
                            raw_input=raw_pos_input,
                        ).to_html(full_html=False, include_plotlyjs='cdn')
                    )

                    if has_numeric_data(actual_df, 'RMS current') and has_numeric_data(avg_df, 'RMS current'):
                        fig_parts.append(
                            pg.plot_motor_app_feedback(
                                "RMS Current",
                                raw_pairs(actual_df, 'RMS current'),
                                raw_pairs(avg_df, 'RMS current'),
                                raw_input=raw_pos_input,
                            ).to_html(full_html=False, include_plotlyjs=False)
                        )
                    else:
                        print("[INFO] Skipping Steering App RMS Current plot: no numeric 'RMS current' data.")

                if fig_parts:
                    result_sections.append({"title": title, "figs": fig_parts})

            except Exception as e:
                print(f"[WARNING] Skipping {sheet} report section: {e}")

        return result_sections
    
    def check_calibration_status(self, dp):
        try:
            actual_df = pd.read_excel(self.results_excel_path, sheet_name='traction_test')
            actual_current = dp.min_max_avg(actual_df, 'current A')
            df_actual = pd.DataFrame(actual_current, columns=['setpoint', 'min', 'max', 'avg'])

            df_actual = df_actual.dropna(subset=['min', 'max'])

            # Calibration check: if any interval exceeds the threshold
            if ((df_actual['max'] - df_actual['min']) > 0.2).any():
                return "not calibrated"
            return "ok"
        except ValueError:
            print("[INFO] Traction calibration status unavailable: sheet 'traction_test' was not produced in this run.")
            return "not available"
        except (FileNotFoundError, KeyError) as exc:
            print(f"[INFO] Traction calibration status unavailable: {exc}")
            return "not available"
        
    
    # Get metadata
    def get_metadata(self,dp):
        """
        Extracts metadata from 'general_traction' and 'general_steering' sheets separately.
        Returns a dict with keys 'traction' and 'steering'.
        """
        meta = {
            'traction': {},
            'steering': {}
        }
        for sheet, key in [('general_traction', 'traction'), ('general_steering', 'steering')]:
            try:
                df = pd.read_excel(self.results_excel_path, sheet_name=sheet, engine='openpyxl')
                for _, row in df.iterrows():
                    param = str(row.get('Parameters', row.get('Parameter', ''))).strip()
                    value = row.get('Controller', row.get('Value', ''))
                    # flatten dict cell if needed
                    if isinstance(value, dict) and value:
                        value = list(value.values())[0]
                    meta[key][param] = value
            except (ValueError, FileNotFoundError, KeyError):
                # sheet not present or invalid, leave empty
                continue

        status = self.check_calibration_status(dp)
        meta['traction']['Threshold Status'] = status

        return meta
    
    @staticmethod
    def get_unique_filename(filepath):
        """
        If filepath exists, append _copy, _copy2, etc. before the extension.
        """
        base, ext = os.path.splitext(filepath)
        counter = 1
        new_filepath = filepath
        while os.path.exists(new_filepath):
            new_filepath = f"{base}_copy{counter}{ext}"
            counter += 1
        return new_filepath

        
    # Rendering plots in HTML tabs
    def render_tabs_html(self, unit_serial_number, sections=None, metadata=None):
        """
        Renders plot sections into an HTML report with tabs,
        showing traction metadata on the left and steering on the right (only if exists).
        """
        if metadata is None:
            metadata = sections or {}
            sections = unit_serial_number or []
            unit_serial_number = Path(self.results_excel_path).stem
            if unit_serial_number.endswith("_test"):
                unit_serial_number = unit_serial_number[:-5]

        plot_name = f"{unit_serial_number}_plots_tabs.html"
        raw_path = os.path.join(self.SAVE_RESULTS_PATH, plot_name)
        plot_path = self.get_unique_filename(raw_path)

        def _render_meta_div(meta_dict):
            if not meta_dict:
                return ''
            html = ['<ul>']
            for k, v in meta_dict.items():
                html.append(f"<li><b>{k}:</b> {v}</li>")
            html.append('</ul>')
            return '\n'.join(html)

        traction_meta = metadata.get('traction', {})
        steering_meta = metadata.get('steering', {})

        with open(plot_path, 'w', encoding='utf-8') as f:
            f.write("<html><head><title>Test Report</title>")
            f.write("""
            <style>
                .metadata-container { display: flex; justify-content: space-between; gap: 40px; }
                .metadata-box { width: 48%; }
                .tab { display: none; }
                .tab.active { display: block; }
                .nav { margin-bottom: 10px; }
                .nav button { margin-right: 5px; padding: 6px 12px; }
            </style>
            <script>
                function showTab(id) {
                    document.querySelectorAll('.tab').forEach(el => el.classList.remove('active'));
                    document.getElementById(id)?.classList.add('active');
                }
                window.onload = function() { showTab('tab0'); };
            </script>
            """)
            f.write("</head><body>")
            f.write(f"<h1>EOL Report: {unit_serial_number}</h1>")

            # Metadata layout
            f.write("<div class='metadata-container'>")
            f.write("<div class='metadata-box'><h2>Traction Metadata</h2>")
            f.write(_render_meta_div(traction_meta))
            f.write("</div>")
            if steering_meta:
                f.write("<div class='metadata-box'><h2>Steering Metadata</h2>")
                f.write(_render_meta_div(steering_meta))
                f.write("</div>")
            f.write("</div><hr>")

            # Tab buttons
            f.write("<div class='nav'>")
            valid = []
            for idx, sec in enumerate(sections):
                if sec and sec.get('title') and sec.get('figs'):
                    valid.append(sec)
                    f.write(f"<button onclick=\"showTab('tab{idx}')\">{sec['title']}</button>")
            f.write("</div>")

            # Tab contents
            for idx, sec in enumerate(valid):
                f.write(f"<div id='tab{idx}' class='tab'>")
                for fig in sec['figs']:
                    if fig:
                        f.write(fig + '<hr>')
                f.write("</div>")

            f.write("</body></html>")

        try:
            if hasattr(os, "startfile"):
                os.startfile(plot_path)
            else:
                import webbrowser

                webbrowser.open(Path(plot_path).resolve().as_uri())
        except Exception as exc:
            print(f"[WARNING] Could not open HTML report automatically: {exc}")
        return plot_path

import os
import pandas as pd
from openpyxl import load_workbook

def is_empty_or_initialized_xlsx(filepath):
    """
    Check if an Excel file is 'empty' (only contains an 'init' sheet with 'initialized', or no real data).
    """
    try:
        wb = load_workbook(filepath, read_only=True)
        sheetnames = wb.sheetnames
        if len(sheetnames) == 0:
            return True
        # Check if the first cell of the first sheet is 'initialized'
        first_sheet = wb[sheetnames[0]]
        first_row = next(first_sheet.iter_rows(min_row=1, max_row=1, values_only=True))
        if str(first_row[0]).strip().lower() == "initialized":
            return True
        # If only one sheet and it's called 'init'
        if len(sheetnames) == 1 and sheetnames[0].lower() == "init":
            return True
        # If the first sheet is truly empty (no data at all)
        df = pd.read_excel(filepath, sheet_name=sheetnames[0])
        if df.empty:
            return True
    except Exception as e:
        print(f"[Warning] Could not open {filepath}: {e}")
        return True  # Treat unreadable files as empty
    return False

def preview_empty_excels(folder):
    """
    List all Excel files in 'folder' that match the empty/initialized pattern.
    """
    to_remove = []
    for fname in os.listdir(folder):
        if fname.endswith(".xlsx"):
            full_path = os.path.join(folder, fname)
            if is_empty_or_initialized_xlsx(full_path):
                to_remove.append(fname)
    if not to_remove:
        print("No empty/initialized Excel files found.")
    else:
        print("These files will be removed:")
        for f in to_remove:
            print(" -", f)
    return to_remove

def delete_files(folder, filelist):
    """
    Delete all files in filelist from the specified folder.
    """
    for fname in filelist:
        try:
            os.remove(os.path.join(folder, fname))
            print(f"Removed: {fname}")
        except Exception as e:
            print(f"Could not remove {fname}: {e}")

if __name__ == "__main__":

    results_folder = r"C:\Users\DaniilYegarmin\Mobotic GmbH\Mobotic - Manufacturing\01_Products\DD-RL-500-225-54SB-48-C-6\DD-RL-500-225-54SB-48-C-6\04_EOL_Results"

    files_to_remove = preview_empty_excels(results_folder)
    if files_to_remove:
        answer = input("Delete these files? (y/n): ").strip().lower()
        if answer == 'y':
            delete_files(results_folder, files_to_remove)
        else:
            print("No files deleted.")
